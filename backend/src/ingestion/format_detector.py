from pathlib import Path

from src.models.document import DocumentFormat


class FormatDetector:
    _PDF_SIGNATURE = b"%PDF"
    _DOCX_SIGNATURE = b"PK\x03\x04"
    _XLSX_SIGNATURE = b"PK\x03\x04"
    _IMAGE_SIGNATURES = {
        b"\x89PNG\r\n\x1a\n": ".png",
        b"\xff\xd8\xff": ".jpg",
        b"II*\x00": ".tiff",
        b"MM\x00*": ".tiff",
    }

    def detect(self, file_path: str | Path) -> DocumentFormat:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = file_path.suffix.lower()

        # Check extension-based detection first for complex formats
        if ext == ".pdf":
            return self._classify_pdf(file_path)
        if ext == ".docx":
            return self._verify_docx(file_path)
        if ext == ".xlsx":
            return self._verify_xlsx(file_path)
        if ext in (".png", ".jpg", ".jpeg", ".tiff", ".tif"):
            return DocumentFormat.IMAGE

        # Fallback to signature-based detection
        return self._detect_by_signature(file_path)

    def get_page_count(self, file_path: str | Path) -> int:
        file_path = Path(file_path)
        fmt = self.detect(file_path)

        if fmt == DocumentFormat.PDF_DIGITAL or fmt == DocumentFormat.PDF_SCANNED:
            return self._pdf_page_count(file_path)
        if fmt == DocumentFormat.DOCX:
            return self._docx_page_count(file_path)
        if fmt == DocumentFormat.XLSX:
            return self._xlsx_sheet_count(file_path)
        if fmt == DocumentFormat.IMAGE:
            return 1
        return 0

    def _classify_pdf(self, file_path: Path) -> DocumentFormat:
        is_scanned = self._is_scanned_pdf(file_path)
        return DocumentFormat.PDF_SCANNED if is_scanned else DocumentFormat.PDF_DIGITAL

    def _is_scanned_pdf(self, file_path: Path) -> bool:
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                total_text = ""
                for page in pdf.pages[:5]:
                    t = page.extract_text()
                    if t:
                        total_text += t
                # If less than 100 chars across first 5 pages, likely scanned
                return len(total_text.strip()) < 100
        except Exception:
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(str(file_path))
                text = "".join(
                    reader.pages[i].extract_text() or ""
                    for i in range(min(5, len(reader.pages)))
                )
                return len(text.strip()) < 100
            except Exception:
                pass
        return True

    def _verify_docx(self, file_path: Path) -> DocumentFormat:
        try:
            from docx import Document as DocxDocument  # noqa: F811
            DocxDocument(str(file_path))
            return DocumentFormat.DOCX
        except Exception:
            return DocumentFormat.IMAGE

    def _verify_xlsx(self, file_path: Path) -> DocumentFormat:
        try:
            from openpyxl import load_workbook
            load_workbook(str(file_path), read_only=True)
            return DocumentFormat.XLSX
        except Exception:
            return DocumentFormat.IMAGE

    def _detect_by_signature(self, file_path: Path) -> DocumentFormat:
        with open(file_path, "rb") as f:
            header = f.read(8)
        if header.startswith(self._PDF_SIGNATURE):
            return self._is_scanned_pdf(file_path) and DocumentFormat.PDF_SCANNED or DocumentFormat.PDF_DIGITAL
        for signature, _ in self._IMAGE_SIGNATURES.items():
            if header.startswith(signature):
                return DocumentFormat.IMAGE
        return DocumentFormat.IMAGE

    def _pdf_page_count(self, file_path: Path) -> int:
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                return len(pdf.pages)
        except Exception:
            return 0

    def _docx_page_count(self, file_path: Path) -> int:
        return 0

    def _xlsx_sheet_count(self, file_path: Path) -> int:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True)
            return len(wb.sheetnames)
        except Exception:
            return 0